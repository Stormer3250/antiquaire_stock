"""Import de fichiers .xlsx/.csv : inspection, correspondance de colonnes, application."""

import csv
import io
import json
import time
import unicodedata
import uuid

import openpyxl
import openpyxl.styles
import openpyxl.utils
from fastapi import APIRouter, HTTPException, Response, UploadFile

from antiquaire import stock
from antiquaire.api import Conn

router = APIRouter()

FIELDS = {"nom", "marque", "categorie", "volume", "degre", "achat", "stock", "fournisseur"}
TOKEN_TTL = 15 * 60
_cache: dict[str, dict] = {}  # token -> {rows, filename, ts}


def normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.casefold().split())


def parse_number(v) -> float:
    if isinstance(v, int | float):
        return float(v)
    n = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    return float(n)


def _is_number(v) -> bool:
    try:
        parse_number(v)
        return True
    except (ValueError, TypeError):
        return False


def _read_rows(filename: str, data: bytes) -> list[list[str]]:
    if filename.lower().endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        rows = [
            ["" if c is None else c for c in row] for row in wb.active.iter_rows(values_only=True)
        ]
    else:
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1")
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        rows = [list(r) for r in csv.reader(io.StringIO(text), dialect)]
    return [r for r in rows if any(str(c).strip() for c in r)]


def _letter(i: int) -> str:
    out = ""
    i += 1
    while i:
        i, rem = divmod(i - 1, 26)
        out = chr(65 + rem) + out
    return out


def inspect_file(filename: str, data: bytes) -> dict:
    rows = _read_rows(filename, data)
    if not rows:
        raise HTTPException(422, "fichier vide ou illisible")
    # en-tête = 1re ligne sans nombre alors que la suite en contient,
    # OU 1re ligne dont ≥2 cellules ressemblent à des noms de champs connus
    known = (
        "nom",
        "marque",
        "categorie",
        "volume",
        "degre",
        "achat",
        "prix",
        "stock",
        "quantite",
        "fournisseur",
        "cond",
        "tva",
    )
    keyword_hits = sum(1 for c in rows[0] if any(k in normalize(c) for k in known))
    has_header = len(rows) > 1 and (
        (
            not any(_is_number(c) for c in rows[0])
            and any(_is_number(c) for r in rows[1:3] for c in r)
        )
        or keyword_hits >= 2
    )
    header = rows[0] if has_header else []
    body = rows[1:] if has_header else rows
    width = max(len(r) for r in body)
    body = [[*r, *[""] * (width - len(r))] for r in body]

    now = time.time()
    for tok in [t for t, v in _cache.items() if now - v["ts"] > TOKEN_TTL]:
        _cache.pop(tok, None)
    token = uuid.uuid4().hex
    _cache[token] = {
        "rows": body,
        "filename": filename,
        "ts": now,
        "offset": 2 if has_header else 1,  # numéros de ligne tels que dans le fichier
    }

    return {
        "token": token,
        "row_count": len(body),
        "columns": [
            {
                "key": str(i),
                "letter": _letter(i),
                "header": str(header[i]) if i < len(header) and str(header[i]).strip() else "",
                "sample": str(body[0][i]) if body else "",
            }
            for i in range(width)
        ],
        "preview": [[str(c) for c in r] for r in body[:4]],
    }


def apply_import(
    conn,
    token: str,
    mapping: dict,
    location_id: int,
    categorie_id: int,
    create_categories: bool = False,
) -> dict:
    entry = _cache.pop(token, None)
    if entry is None or time.time() - entry["ts"] > TOKEN_TTL:
        raise HTTPException(410, "import expiré, redéposez le fichier")
    fields_by_col = {int(k): v for k, v in mapping.items() if v in FIELDS}
    if "nom" not in fields_by_col.values():
        raise HTTPException(422, "la colonne du nom est obligatoire")

    cats = {
        normalize(r["nom"]): r["id"]
        for r in conn.execute("SELECT id, nom FROM categories WHERE active = 1")
    }
    existing = {
        normalize(r["nom"]): r["id"]
        for r in conn.execute("SELECT id, nom FROM refs WHERE active = 1")
    }
    numeric = {"volume": "vol_cl", "degre": "abv", "achat": "achat_ht"}
    texts = {"nom": "nom", "marque": "marque", "fournisseur": "fournisseur"}

    created = updated = 0
    errors: list[str] = []
    comptages: list[dict] = []
    seen_fournisseurs: list[str] = []
    try:
        for line_no, row in enumerate(entry["rows"], start=entry["offset"]):
            values = {field: row[col] for col, field in fields_by_col.items() if col < len(row)}
            nom = str(values.get("nom", "")).strip()
            if not nom:
                errors.append(f"ligne {line_no}: nom vide, ignorée")
                continue
            patch: dict = {}
            bad = None
            for field, col_name in numeric.items():
                if field in values and str(values[field]).strip() != "":
                    try:
                        patch[col_name] = parse_number(values[field])
                    except (ValueError, TypeError):
                        bad = f"ligne {line_no}: {field} illisible « {values[field]} »"
            stock_qty = None
            if "stock" in values and str(values["stock"]).strip() != "":
                try:
                    stock_qty = parse_number(values["stock"])
                except (ValueError, TypeError):
                    bad = f"ligne {line_no}: stock illisible « {values['stock']} »"
            if "categorie" in values and str(values["categorie"]).strip():
                cat_nom = str(values["categorie"]).strip()
                cat_id = cats.get(normalize(cat_nom))
                if cat_id is None and create_categories:
                    cur = conn.execute(
                        "INSERT INTO categories (nom, position) VALUES (?,"
                        " (SELECT COALESCE(MAX(position), 0) + 1 FROM categories))",
                        (cat_nom,),
                    )
                    cat_id = cur.lastrowid
                    cats[normalize(cat_nom)] = cat_id
                if cat_id is None:
                    bad = f"ligne {line_no}: catégorie inconnue « {values['categorie']} »"
                else:
                    patch["categorie_id"] = cat_id
            if bad:
                errors.append(bad)
                continue
            for field, col_name in texts.items():
                if field in values and str(values[field]).strip():
                    patch[col_name] = str(values[field]).strip()
            if patch.get("fournisseur"):
                seen_fournisseurs.append(patch["fournisseur"])

            ref_id = existing.get(normalize(nom))
            if ref_id is None:
                patch.setdefault("categorie_id", categorie_id)
                patch.setdefault("seuil", 2)
                patch["nom"] = nom
                cols = ", ".join(patch)
                marks = ", ".join("?" for _ in patch)
                cur = conn.execute(
                    f"INSERT INTO refs ({cols}, created_at) VALUES ({marks}, ?)",
                    [*patch.values(), stock.now()],
                )
                ref_id = cur.lastrowid
                existing[normalize(nom)] = ref_id
                created += 1
            else:
                if patch:
                    sets = ", ".join(f"{k} = ?" for k in patch)
                    conn.execute(f"UPDATE refs SET {sets} WHERE id = ?", [*patch.values(), ref_id])
                updated += 1
            if stock_qty is not None:
                comptages.append(
                    {
                        "ref_id": ref_id,
                        "location_id": location_id,
                        "type": "comptage",
                        "quantity": stock_qty,
                        "note": f"import {entry['filename']}",
                    }
                )
        for line in comptages:
            conn.execute(
                "INSERT INTO movements (ref_id, location_id, type, quantity, source, note,"
                " created_at) VALUES (?, ?, ?, ?, 'import', ?, ?)",
                (
                    line["ref_id"],
                    line["location_id"],
                    line["type"],
                    line["quantity"],
                    line["note"],
                    stock.now(),
                ),
            )
        # les fournisseurs jamais vus rejoignent la liste de Configuration
        if seen_fournisseurs:
            lists = json.loads(
                conn.execute("SELECT value FROM settings WHERE key = 'lists'").fetchone()[0]
            )
            known = {normalize(f) for f in lists["fournisseurs"]}
            for f in seen_fournisseurs:
                if normalize(f) not in known:
                    lists["fournisseurs"].append(f)
                    known.add(normalize(f))
            conn.execute("UPDATE settings SET value = ? WHERE key = 'lists'", (json.dumps(lists),))
        conn.execute(
            "INSERT INTO imports (filename, line_count, created_count, updated_count, created_at)"
            " VALUES (?, ?, ?, ?, ?)",
            (entry["filename"], len(entry["rows"]), created, updated, stock.now()),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {"created": created, "updated": updated, "lines": len(entry["rows"]), "errors": errors}


# ---------- modèle à télécharger ----------

# en-têtes choisis pour être reconnus tels quels par le pré-mappage de l'interface
TEMPLATE_HEADERS = [
    "Nom",
    "Marque",
    "Catégorie",
    "Volume (cl)",
    "Degré alcoolique",
    "Prix d'achat HT",
    "Quantité en stock",
    "Fournisseur",
]
TEMPLATE_ROWS = [
    ["Gin London Dry", "Sipsmith · Londres", "Spiritueux", 70, 41.6, 18.40, 6, "Dugas"],
    ["Chenin sec Loire", "Domaine Huet", "Vin", 75, 13, 16.50, 9, "Vinifera"],
]


@router.get("/import/template")
def import_template(format: str = "xlsx"):
    if format == "csv":
        out = io.StringIO()
        writer = csv.writer(out, delimiter=";")
        writer.writerow(TEMPLATE_HEADERS)
        writer.writerows(TEMPLATE_ROWS)
        return Response(
            out.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="modele-antiquaire.csv"'},
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Références"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in TEMPLATE_ROWS:
        ws.append(row)
    for i, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(header) + 4, 14)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modele-antiquaire.xlsx"'},
    )


# ---------- routes ----------


@router.post("/import/inspect")
def import_inspect(file: UploadFile):
    return inspect_file(file.filename or "fichier", file.file.read())


@router.post("/import/apply")
def import_apply(conn: Conn, body: dict):
    for key in ("token", "mapping", "location_id", "categorie_id"):
        if not body.get(key):
            raise HTTPException(422, f"{key} obligatoire")
    return apply_import(
        conn,
        body["token"],
        body["mapping"],
        body["location_id"],
        body["categorie_id"],
        create_categories=bool(body.get("create_categories")),
    )


@router.get("/imports")
def imports_history(conn: Conn):
    rows = conn.execute("SELECT * FROM imports ORDER BY id DESC LIMIT 20").fetchall()
    return {"imports": [dict(r) for r in rows]}
